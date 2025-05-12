import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Literal, Optional
from datetime import datetime
from .distributor_data import get_distributor_info
from .tabs.costs_data import load_costs_sheet
from .tabs.tusd_or_te_market_data import load_tusd_or_te_market_sheet
from .tabs.tusd_or_te_data import load_tusd_or_te_sheet, TusdOrTe, create_mixed_tusd_or_te_worksheet
from .tabs.effect_data import load_effect_sheet
from .tabs.reh_tables_data import load_reh_tables_sheet
from .helper import join_sheets_vertically, get_date_from


def process_and_merge_workbook(
    uploaded_file_path: str,
    acronym: str,
    tariff_process: Literal["Ajuste EER ANGRA III", "Liminar abrace", "Reajuste", "Revisão", "Revisão Extraordinária", "Tarifas Iniciais"],
    process_date_str: str
):
    uploaded_workbook = load_workbook(uploaded_file_path, data_only=True)
    process_date = get_date_from(process_date_str)

    new_workbook = _filtered_workbook(
        workbook=uploaded_workbook,
        acronym=acronym,
        tariff_process=tariff_process,
        process_date=process_date
    )

    workbooks_to_merge = [new_workbook]

    db_path = os.path.join(os.path.dirname(__file__), "../storage/banco.xlsx")
    db_path = os.path.abspath(db_path)

    if os.path.exists(db_path):
        existing_workbook = load_workbook(db_path, data_only=True)
        workbooks_to_merge.append(existing_workbook)

    _join_sheets(
        workbooks=workbooks_to_merge,
        output_name=db_path
    )


def _filtered_workbook(
    workbook: Workbook,
    acronym: str, 
    tariff_process: Literal["Ajuste EER ANGRA III", "Liminar abrace", "Reajuste", "Revisão", "Revisão Extraordinária", "Tarifas Iniciais"],
    process_date = datetime
) -> Workbook:
    distributor_info = get_distributor_info(acronym=acronym)
    distributor_info = {
        'Nome': distributor_info['name'],
        'Sigla': acronym,
        'Concessionária/Permissionária': distributor_info['agent'],
        'Código da Empresa': distributor_info['company_code'],
        'ID Agente': distributor_info['agent_id'],
        'ID Concessão': distributor_info['concession_id'],
        'Processo Tarifário': tariff_process,
        'Data do processo tarifário em processamento': process_date
    }

    distributor_header = list(distributor_info.keys())

    new_workbook = Workbook()
    default_sheet = new_workbook.active
    new_workbook.remove(default_sheet)

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_costs_sheet(workbook=workbook),
        tab_name='CUSTOS'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_market_sheet(workbook=workbook, tusd_or_te="TUSD"),
        tab_name='MERCADO TUSD'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_sheet(workbook=workbook, tusd_or_te=TusdOrTe.TUSD),
        tab_name='TUSD',
        hide_first_line=True
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_market_sheet(workbook=workbook, tusd_or_te="TE"),
        tab_name='MERCADO TE'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_sheet(workbook=workbook, tusd_or_te=TusdOrTe.TE),
        tab_name='TE',
        hide_first_line=True
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_effect_sheet(workbook=workbook),
        tab_name='EFEITO'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_reh_tables_sheet(workbook=workbook),
        tab_name="TABELAS REH"
    )

    if len(new_workbook.sheetnames) == 0:
        new_workbook.create_sheet(title="Sheet")

    return new_workbook


def _create_db_tab(distributor_info: dict[str, any], distributor_header: list[str], workbook: Workbook, worksheet: Optional[Worksheet], tab_name: str, hide_first_line: bool = False):
    if not worksheet:
        return

    new_worksheet = workbook.create_sheet(title=tab_name)
    worksheet_header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    is_worksheet_empty = all(cell is None or str(cell).strip() == "" for cell in worksheet_header)

    if is_worksheet_empty:
        return

    new_worksheet.append(distributor_header + worksheet_header)

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    counter = 0

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
        if not all(cell is None for cell in row):
            if hide_first_line and counter == 0:
                length = len(distributor_info.values())
                empty_row = []

                for _ in range(length):
                    empty_row.append("")

                new_row = empty_row + list(row)
                new_worksheet.append(new_row)
                counter += 1

                continue

            new_row = list(distributor_info.values()) + list(row)
            new_worksheet.append(new_row)


def _join_sheets(workbooks: list[Workbook], output_name: str):
    output_workbook = Workbook()
    del output_workbook["Sheet"]

    _create_mixed_worksheet(
        workbooks=workbooks,
        tab_name="CUSTOS",
        output_workbook=output_workbook
    )

    _create_mixed_worksheet(
        workbooks=workbooks,
        tab_name="MERCADO TUSD",
        output_workbook=output_workbook
    )

    create_mixed_tusd_or_te_worksheet(
        workbooks=workbooks,
        tusd_or_te=TusdOrTe.TUSD,
        output_workbook=output_workbook
    )

    _create_mixed_worksheet(
        workbooks=workbooks,
        tab_name="MERCADO TE",
        output_workbook=output_workbook
    )

    create_mixed_tusd_or_te_worksheet(
        workbooks=workbooks,
        tusd_or_te=TusdOrTe.TE,
        output_workbook=output_workbook
    )

    _create_mixed_worksheet(
        workbooks=workbooks,
        tab_name="EFEITO",
        output_workbook=output_workbook
    )

    _create_mixed_worksheet(
        workbooks=workbooks,
        tab_name="TABELAS REH",
        output_workbook=output_workbook
    )

    output_workbook.save(output_name)


def _create_mixed_worksheet(workbooks: list[Workbook], tab_name: str, output_workbook: Workbook):
    worksheets = [
        workbook[tab_name]
        for workbook in workbooks
        if tab_name in workbook.sheetnames and
        any(row for row in workbook[tab_name].iter_rows(values_only=True) if any(cell is not None for cell in row))
    ]

    mixed_worksheet = join_sheets_vertically(worksheets)
    new_worksheet = output_workbook.create_sheet(title=tab_name)

    for row in mixed_worksheet.iter_rows(values_only=True):
        new_worksheet.append(row)